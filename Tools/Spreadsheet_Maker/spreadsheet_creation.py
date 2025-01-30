import pandas as pd
from typing import List, Dict, Optional
from pathlib import Path
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from google.oauth2.credentials import Credentials
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google.oauth2 import service_account
from googleapiclient.discovery import build
import os
from dotenv import load_dotenv
import json

# Load environment variables
load_dotenv()

"""
IMPORTANT: Before running this script, make sure to activate the work environment:
$ src work

This ensures the correct Python environment and dependencies are available.
"""


def style_excel_headers(
    worksheet,
    num_columns: int,
    header_fill_color: str = "CCCCCC",
    header_font_bold: bool = True
) -> None:
    """
    Apply styling to the header row of an Excel worksheet.

    Args:
        worksheet: An openpyxl worksheet object.
        num_columns: Number of columns that have headers.
        header_fill_color: HEX color code for the header background.
        header_font_bold: Whether the header font should be bold.
    """
    header_fill = PatternFill(start_color=header_fill_color, end_color=header_fill_color, fill_type="solid")
    header_font = Font(bold=header_font_bold)

    # Assuming headers are in the first row
    for col_idx in range(1, num_columns + 1):
        cell = worksheet.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font


def adjust_column_widths(
    worksheet,
    df: pd.DataFrame,
    max_col_width: int = 50
) -> None:
    """
    Adjust column widths in an Excel worksheet based on content length.

    Args:
        worksheet: An openpyxl worksheet object.
        df: The DataFrame that was written to the worksheet.
        max_col_width: The maximum allowed width for any column.
    """
    for idx, column in enumerate(df.columns, start=1):
        # Determine max length of data plus header
        max_length = max(df[column].astype(str).apply(len).max(), len(str(column)))
        adjusted_width = min(max_length + 2, max_col_width)
        worksheet.column_dimensions[get_column_letter(idx)].width = adjusted_width


def validate_input_data(
    data: List[Dict],
    columns: Optional[List[str]]
) -> List[str]:
    """
    Validate the input data and determine the final columns to use.

    Args:
        data: List of dictionaries containing the data.
        columns: Optional list of column names to include.

    Returns:
        List[str]: The columns that will be used in the DataFrame.

    Raises:
        ValueError: If the data is invalid or required columns are missing.
    """
    if not data or not isinstance(data, list):
        raise ValueError("Input data must be a non-empty list of dictionaries.")
    if not all(isinstance(row, dict) for row in data):
        raise ValueError("All items in 'data' must be dictionaries.")

    # Auto-determine columns if not explicitly provided
    if columns is None:
        columns = list(data[0].keys())

    # Check for missing columns
    for col in columns:
        if any(col not in row for row in data):
            raise ValueError(f"Missing column '{col}' in one or more rows of the input data.")

    return columns


def create_spreadsheet_from_data(
    data: List[Dict],
    columns: Optional[List[str]] = None,
    filename: str = 'output.xlsx',
    sheet_name: str = 'Sheet1',
    style_headers_flag: bool = True,
    header_fill_color: str = "CCCCCC",
    header_font_bold: bool = True,
    max_col_width: int = 50
) -> str:
    """
    Creates an Excel spreadsheet from structured input data.

    This function takes a list of dictionaries, converts it to a DataFrame, and writes it to an Excel file.
    It provides options to style headers and adjust column widths. This modular design is intended for 
    integration with LLMs, allowing for easy spreadsheet generation from dynamically obtained data.

    Args:
        data: A list of dictionaries, where each dict represents a row of data.
        columns: A list of column names to include in the sheet. If None, columns will be inferred.
        filename: The name or path of the output Excel file.
        sheet_name: The name of the sheet to create or overwrite.
        style_headers_flag: Whether to style the header row.
        header_fill_color: Background color for the header row cells.
        header_font_bold: Whether the header fonts should be bold.
        max_col_width: Maximum width for auto-adjusted columns.

    Returns:
        str: The absolute path to the created Excel file.

    Raises:
        ValueError: If the input data is invalid or required columns are missing.
    """
    # Validate and determine columns
    columns = validate_input_data(data, columns)

    # Create the DataFrame
    df = pd.DataFrame(data, columns=columns)

    # Write the DataFrame to an Excel file
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)

        # Access the workbook and worksheet
        worksheet = writer.sheets[sheet_name]

        # Adjust column widths
        adjust_column_widths(worksheet, df, max_col_width=max_col_width)

        # Apply header styling if requested
        if style_headers_flag:
            style_excel_headers(
                worksheet,
                num_columns=len(df.columns),
                header_fill_color=header_fill_color,
                header_font_bold=header_font_bold
            )

    return str(Path(filename).resolve())


def create_google_sheet_from_data(
    data: List[Dict],
    columns: Optional[List[str]] = None,
    sheet_name: str = 'Sheet1',
    credentials_path: Optional[str] = None,
) -> str:
    """
    Creates a Google Sheet from structured input data.

    Args:
        data: A list of dictionaries, where each dict represents a row of data.
        columns: A list of column names to include in the sheet. If None, columns will be inferred.
        sheet_name: The name of the sheet to create.
        credentials_path: Not used anymore, kept for backwards compatibility.

    Returns:
        str: The URL of the created Google Sheet.

    Raises:
        ValueError: If the input data is invalid or required columns are missing.
        Exception: If there are issues with Google Sheets API authentication or creation.
    """
    # Validate and determine columns
    columns = validate_input_data(data, columns)

    # Create the DataFrame
    df = pd.DataFrame(data, columns=columns)

    # Get credentials from environment variables
    creds_info = {
        "account": "",
        "client_id": os.getenv("GOOGLE_CLIENT_ID"),
        "client_secret": os.getenv("GOOGLE_CLIENT_SECRET"),
        "quota_project_id": os.getenv("GOOGLE_QUOTA_PROJECT_ID"),
        "refresh_token": os.getenv("GOOGLE_REFRESH_TOKEN"),
        "type": "authorized_user",
        "universe_domain": os.getenv("GOOGLE_UNIVERSE_DOMAIN")
    }

    # Create credentials from the OAuth2 info
    credentials = Credentials.from_authorized_user_info(
        info=creds_info,
        scopes=['https://www.googleapis.com/auth/spreadsheets']
    )

    if not credentials or not credentials.valid:
        if credentials and credentials.expired and credentials.refresh_token:
            credentials.refresh(Request())

    # Create Google Sheets API service
    service = build('sheets', 'v4', credentials=credentials)

    # Create a new spreadsheet
    spreadsheet = service.spreadsheets().create(body={
        'properties': {'title': sheet_name}
    }).execute()
    spreadsheet_id = spreadsheet['spreadsheetId']

    # Prepare the values for the sheet
    values = [columns]  # Header row
    values.extend(df.values.tolist())  # Data rows

    # Update the values
    body = {
        'values': values
    }
    service.spreadsheets().values().update(
        spreadsheetId=spreadsheet_id,
        range=f'{sheet_name}!A1',
        valueInputOption='RAW',
        body=body
    ).execute()

    # Format header row
    requests = [{
        'repeatCell': {
            'range': {
                'sheetId': 0,
                'startRowIndex': 0,
                'endRowIndex': 1
            },
            'cell': {
                'userEnteredFormat': {
                    'backgroundColor': {
                        'red': 0.8,
                        'green': 0.8,
                        'blue': 0.8,
                        'alpha': 1
                    },
                    'textFormat': {
                        'bold': True
                    }
                }
            },
            'fields': 'userEnteredFormat(backgroundColor,textFormat)'
        }
    }]

    # Auto-resize columns
    requests.append({
        'autoResizeDimensions': {
            'dimensions': {
                'sheetId': 0,
                'dimension': 'COLUMNS',
                'startIndex': 0,
                'endIndex': len(columns)
            }
        }
    })

    # Apply formatting
    service.spreadsheets().batchUpdate(
        spreadsheetId=spreadsheet_id,
        body={'requests': requests}
    ).execute()

    # Return the spreadsheet URL
    return f'https://docs.google.com/spreadsheets/d/{spreadsheet_id}'


# Example usage:
if __name__ == "__main__":
    sample_data = [
        {"id": 1, "name": "John Doe", "email": "john@example.com", "age": 30},
        {"id": 2, "name": "Jane Smith", "email": "jane@example.com", "age": 25},
        {"id": 3, "name": "Bob Johnson", "email": "bob@example.com", "age": 35}
    ]

    sheet_url = create_google_sheet_from_data(
        data=sample_data,
        sheet_name='User Data'
    )
    print(f"Spreadsheet created at: {sheet_url}")